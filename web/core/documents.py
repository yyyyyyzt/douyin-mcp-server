"""上传文档解析：支持 PDF / Excel 报价单等，提取纯文本供问答与合同审查。"""

import io
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

MAX_FILE_BYTES = 10 * 1024 * 1024  # 10MB
MAX_TEXT_CHARS = 30_000
ALLOWED_SUFFIXES = {".pdf", ".xlsx", ".xlsm"}


class DocumentParseError(Exception):
    """文档解析失败。"""


def _truncate(text: str) -> tuple[str, bool]:
    text = (text or "").strip()
    if len(text) <= MAX_TEXT_CHARS:
        return text, False
    return text[:MAX_TEXT_CHARS] + "\n\n…（内容过长，已截断）", True


def parse_pdf(content: bytes) -> str:
    """从 PDF 提取文本。"""
    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception as e:
        raise DocumentParseError(f"PDF 读取失败: {e}") from e

    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            pages.append("")
    text = "\n".join(pages).strip()
    if not text:
        raise DocumentParseError("PDF 中未识别到可提取文本（可能是扫描件）")
    return text


def parse_excel(content: bytes) -> str:
    """从 Excel 工作簿提取文本（各 sheet 表格转 TSV 风格）。"""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise DocumentParseError(f"Excel 读取失败: {e}") from e

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"## {sheet_name}\n" + "\n".join(rows))
    wb.close()

    text = "\n\n".join(parts).strip()
    if not text:
        raise DocumentParseError("Excel 中未识别到有效单元格内容")
    return text


def parse_document(filename: str, content: bytes) -> dict:
    """解析上传文件，返回文件名、正文与元信息。"""
    if not filename:
        raise DocumentParseError("文件名为空")
    if len(content) > MAX_FILE_BYTES:
        raise DocumentParseError(f"文件过大，请上传 {MAX_FILE_BYTES // (1024 * 1024)}MB 以内的文件")

    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        raw = parse_pdf(content)
        file_type = "pdf"
    elif suffix in (".xlsx", ".xlsm"):
        raw = parse_excel(content)
        file_type = "excel"
    elif suffix == ".xls":
        raise DocumentParseError("暂不支持旧版 .xls，请另存为 .xlsx 后上传")
    else:
        raise DocumentParseError(f"不支持的文件格式「{suffix}」，请上传 PDF 或 Excel（.xlsx）")

    text, truncated = _truncate(raw)
    if not text:
        raise DocumentParseError("未能从文件中提取到文本")

    return {
        "filename": Path(filename).name,
        "file_type": file_type,
        "text": text,
        "char_count": len(text),
        "truncated": truncated,
    }
